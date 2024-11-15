import FlowRank_General as FR
from sklearn.metrics.cluster import normalized_mutual_info_score as NMI
from sknetwork.ranking import PageRank
from scipy import sparse
import networkx as nx
import numpy as np
import metric as met
from llist import dllist
import matplotlib.pyplot as plt
import igraph as ig
import debug
import Cust_Top as Cust
import leidenalg as la
from matplotlib.backends.backend_pdf import PdfPages
#import deque
from collections import deque


'''
FUnctions for Top k% induced subgraph

'''             
def getInducedSubgraph(G, top_nodes): #G = original graph, k = pick top k percent, node2FR = node to FR value
    #Remove nodes also remove adjacent edges
    H = G.copy()
    for u in G.nodes:
        if u not in top_nodes:
            H.remove_node(u)
    return H

def part_to_full_label(partition, original_n, mapping = None): #partition to labels (Returns full labels)
    label_1=[-1]*(original_n)
    '''
    iGraph has different node numbering than networkx (With mapping)
    '''
    if mapping is not None:
        for i, community in enumerate(partition):
            for node in community:
                label_1[mapping[node]] = i   
    else:    
        c=0
        for sets in partition:
            for ell in sets:
                label_1[ell]=c
            
            c=c+1
    #print(label_compressed)

    return label_1

def FlowRank_Func(edge_list,vlist,walk_len_c1,c_const=0,type=0):
    if type==0:
        return FR.FLOW(edge_list,vlist,walk_len_c1,c_const)
    elif type==1:
        return FR.FLOW_ng(edge_list,vlist,walk_len_c1,c_const)
    elif type==2:
        return FR.FLOW_ng_prop(edge_list,vlist,walk_len_c1,c_const)

def networkX_to_adjMatrix(G):
    # Get a list of nodes and create a mapping to indices
    nodes = list(G.nodes())
    node_indices = {node: idx for idx, node in enumerate(nodes)}

    # Build the adjacency matrix
    row_indices = []
    col_indices = []
    data = []

    for source, target in G.edges():
        row_indices.append(node_indices[source])
        col_indices.append(node_indices[target])
        data.append(1)

    num_nodes = len(nodes)
    adjacency_matrix = sparse.csr_matrix(
        (data, (row_indices, col_indices)), shape=(num_nodes, num_nodes)
    )

    return node_indices, adjacency_matrix

def calc_FlowRank(graph, FR_type, walk_len_c1):
    node2FR = dict()
    if FR_type==4:
        # pg_rank = nx.pagerank(graph,alpha=0.85) #alpha = 0.85 is the default
        # node2FR = {k: pg_rank[k]*graph.number_of_nodes() for k in pg_rank}

        pagerank = PageRank(damping_factor=0.85)
        node_indices, adj_matrix = networkX_to_adjMatrix(graph)
        pagerank.fit(adj_matrix)
        scores = pagerank.scores_
        node2FR = {node: scores[idx] for node, idx in node_indices.items()}
    elif FR_type==3:
        pg_rank = nx.pagerank(graph,alpha=0.85) #alpha = 0.85 is the default
        node2FR = {k: pg_rank[k]*len(graph.nodes()) for k in pg_rank}
    else:
        for i in FlowRank_Func(graph.edges(),graph.nodes(),walk_len_c1,0,FR_type):
            node_num = int(i[1])
            node2FR[node_num] = i[0]
    
    return node2FR

def networkx_to_igraph(G):
    I = ig.Graph(directed=True)
    I.add_vertices(G.number_of_nodes())
    ig_to_nx_idx = list(G.nodes)
    nx_to_ig_idx = {ig_to_nx_idx[i]: i for i in range(len(ig_to_nx_idx))}
    I.vs['name'] = ig_to_nx_idx
    edges = [(nx_to_ig_idx[u], nx_to_ig_idx[v]) for u, v in G.edges]
    I.add_edges(edges)
    return I, ig_to_nx_idx, nx_to_ig_idx


'''
FUnctions for Datasets and Output

'''        
import data_utils_ch as data_util

def data_to_graph(name, survive=0):
    #scRNA datasets
    if name in ['Zhengmix8eq','ALM', 'AMB', 'Baron_Human', 'Baron_Mouse', 'Muraro', 'Segerstolpe', 'Tcell-medicine', 'TM', 'VISP', 'Xin', 'Zheng']:
        edge_list,vlist,n,label=data_util.local_SCRNA(name)
        #print("Dataset names is ",name," |V|, |E| #clusters= ",n,len(edge_list),len(set(label)))

    #These are for the bulk-RNA datasets.
    elif name in ['mRNA','miRNA']:
        #for survive in [0,1]:
        edge_list,vlist,label,n=data_util.local_bulkRNA(name,survive)
        label=data_util.set_labels(label)
        #print("Dataset names is ",name," |V|, |E| #clusters= ",n,len(edge_list),len(set(label)))

    #These contain image and document data.
    #datanames=['FashionMNIST','MNIST','seeds','breast-cancer','Omniglot','bbc_news','20NewsGroups','biorxiv','big_patent']
    elif name in ['FashionMNIST','MNIST','seeds','breast-cancer','Omniglot','bbc_news','20NewsGroups_tfdif','biorxiv','big_patent']:
        edge_list, label=data_util.load_data(name,kchoice=10)
        n=len(label)
        vlist=[i for i in range(n)]
        label=data_util.set_labels(label)
        #print("Dataset names is ",name," |V|, |E| #clusters= ",n,len(edge_list),len(set(label)))


    #These are data of 4 popular directed graphs. 
    elif name in ['Cora','Cora full','Citeseer','Eu core']:
        edge_list,vlist,label,n,good_v=data_util.graph_database(name)
        label=data_util.set_labels(label)
        #print("Dataset names is ",name," |V|, |E| #clusters= ",n,len(edge_list),len(set(label)))
    
    G = nx.DiGraph(edge_list)
    return G, label

def write_out(data, name, max_nmi, max_purity):
    with open('./Results/'+data+'.txt', 'a') as f:
        f.write(name+'\n')
    with open('./Results/'+data+'.txt', 'a') as f:
        f.write('   Max NMI = [' + str(round(max_nmi[0],3)) + ',' + str(round(max_nmi[1],3)) + ']' + ' res:= ' + str(round(max_nmi[2],3)) + '\n')
    
    with open('./Results/'+data+'.txt', 'a') as f:
        f.write('   Partition sizes: [')
    for i in max_nmi[4]:
        #print(len(i),end=' ')
        with open('./Results/'+data+'.txt', 'a') as f:
            f.write(str(len(i))+', ')
    with open('./Results/'+data+'.txt', 'a') as f:
        f.write('] ' + '# of comm: ' + str(len(max_nmi[4]))+ '\n')
    
    
    with open('./Results/'+data+'.txt', 'a') as f:
        f.write('   Max Purity = [' + str(round(max_purity[0],3)) + ',' + str(round(max_purity[1],3)) + ']' + ' res:= ' + str(round(max_purity[2],3)) + '\n')
    with open('./Results/'+data+'.txt', 'a') as f:
        f.write('   Partition sizes: [')
    for i in max_purity[4]:
        with open('./Results/'+data+'.txt', 'a') as f:
            f.write(str(len(i))+', ')
    with open('./Results/'+data+'.txt', 'a') as f:
        f.write('] ' + '# of comm: ' + str(len(max_purity[4]))+ '\n')


'''
Functions for strong majority voting

'''
from collections import Counter, defaultdict
import random
def vote(G, H_label, node):
    #check every outgoing edge of node and counter the majority vote
    if len(G.out_edges(node)) == 0:
        return -1
    vt = defaultdict(int)
    for i in G.out_edges(node):
        #print('i:',i)
        vt[H_label[i[1]]] += 1
    #get the most common label
    most_common_label, most_common_count = Counter(vt).most_common(1)[0]
    
    #Strong Majority Vote
    if most_common_count <= len(G.out_edges(node))/2:
        return -1
    else:
        return most_common_label


def calc_balancedness(selected_labels_dict, cluster_sizes):
    #loop through the selected labels and calculate the balancedness
    min_cluster = 2
    max_cluster = -1
    for cluster, cluster_size in cluster_sizes.items(): #key = cluster #, value = count of nodes
        cnt = selected_labels_dict[cluster]
        ratio = cnt/cluster_size
        #print('cluster:',cluster, 'ratio:',round(ratio,3), end=' /')
        if ratio < min_cluster:
            min_cluster = ratio
        if ratio > max_cluster:
            max_cluster = ratio
    #print('min_cluster:',round(min_cluster,4), 'max_cluster:', round(max_cluster,4))
    return min_cluster/max_cluster

def calc_preservation(selected_labels_dict, cluster_sizes, num_total):
    num_selected = sum(selected_labels_dict.values())

    ratio = 0
    for cluster, cluster_size in cluster_sizes.items(): #key = cluster #, value = count of nodes
        cnt = selected_labels_dict[cluster]
        ratio += min(cnt/cluster_size, num_selected/num_total)
    ratio = (ratio/num_selected)*(num_total/len(cluster_sizes))

    return ratio

def get_compressed_labels(label, H_label, node_list=None):
    label_compressed = []
    H_label_compressed = []
    
    if node_list is None:
        for idx_, label_ in enumerate(H_label):
            if label_ != -1:
                H_label_compressed.append(label_)
                label_compressed.append(label[idx_])
    else:
        for i in node_list:
            if H_label[i] != -1:
                H_label_compressed.append(H_label[i])
                label_compressed.append(label[i])
    return label_compressed, H_label_compressed

def calc_NMI_Purity(label, H_label):
    label_compressed, H_label_compressed = get_compressed_labels(label, H_label)
    NMI_ = NMI(label_compressed, H_label_compressed)
    Purity_ = met.purity_score(label_compressed, H_label_compressed)
    return NMI_, Purity_

def merge_by_vote(top_nodes, nodes_rest, H_label, G, label, selected_labels_dict, cluster_sizes):
    #initial_num_of_nodes = len(top_nodes) - int(k*G.number_of_nodes())
    initial_num_of_nodes = 0
    #print('Initial number of nodes:', initial_num_of_nodes)
    #print('type of initial_num_of_nodes:', type(initial_num_of_nodes))
    flag = 1
    cnt = 0
    n = G.number_of_nodes()
    
    True_label_compressed, H_label_compressed = get_compressed_labels(label, H_label, top_nodes)
    
    NMI_List = [NMI(H_label_compressed, True_label_compressed)]*(initial_num_of_nodes+1)
    Purity_List = [met.purity_score(True_label_compressed,H_label_compressed)]*(initial_num_of_nodes+1)
    Balance_List = [calc_balancedness(selected_labels_dict, cluster_sizes)]*(initial_num_of_nodes+1)
    Preserv_List = [calc_preservation(selected_labels_dict, cluster_sizes, G.number_of_nodes())]*(initial_num_of_nodes+1)

    total_inEdge = 0
    for node in top_nodes:
        if H_label[node] != -1:
            total_inEdge += len(G.in_edges(node))
    InEdge_List = [total_inEdge]
 
    node_linked_list = dllist(nodes_rest)

    
    while(flag):
        flag = 0
        #Reset Traversal (If found a node to merge, start the loop over)    
        nd = node_linked_list.first
        while nd:
            
            node = nd.value
            nd_next = nd.next
            #Already has a community assigned (Skip)
            if H_label[node] != -1:
                nd = nd_next
                continue
            
            new_comm = vote(G, H_label, node)
            if new_comm != -1:
                H_label[node] = new_comm
                node_linked_list.remove(nd)
                H_label_compressed.append(new_comm)
                True_label_compressed.append(label[node])
                InEdge_List.append(InEdge_List[-1] + len(G.in_edges(node)))
                selected_labels_dict[label[node]] += 1
                flag=1
                cnt +=1
                #Calculate new NMI every 5% of nodes
                if cnt > n/20:
                    new_nmi = NMI(H_label_compressed, True_label_compressed)
                    new_purity = met.purity_score(True_label_compressed,H_label_compressed) 
                    NMI_List.append(new_nmi)
                    Purity_List.append(new_purity) 
                    Balance_List.append(calc_balancedness(selected_labels_dict, cluster_sizes))
                    Preserv_List.append(calc_preservation(selected_labels_dict, cluster_sizes, G.number_of_nodes()))
                    #print each entry of selected_labels_dict and cluster_sizes & compare
                    
                    cnt = 0
                else:
                    NMI_List.append(NMI_List[-1])
                    Purity_List.append(Purity_List[-1])
                    Balance_List.append(Balance_List[-1])
                    Preserv_List.append(Preserv_List[-1])
            nd = nd_next
            
        
    NMI_List[-1] = NMI(H_label_compressed, True_label_compressed)
    Purity_List[-1] = met.purity_score(True_label_compressed,H_label_compressed)
    Balance_List[-1] = calc_balancedness(selected_labels_dict, cluster_sizes)
    Preserv_List[-1] = calc_preservation(selected_labels_dict, cluster_sizes, G.number_of_nodes())
    return NMI_List, Purity_List, Balance_List, Preserv_List, InEdge_List

def effective_cluster_accuracy(G, label, selected_labels_dict, res):
    #For each cluster, randomly select nodes for selected_labels_dict[cluster] amount of nodes
    nodes_selected = []
    nodes_in_each_cluster = defaultdict(list)
    for i, node_label in enumerate(label):
        nodes_in_each_cluster[node_label].append(i)
    for cluster, count in selected_labels_dict.items():
        #if cluster == -1, print
        if cluster == -1:
            print('Error: Cluster -1 found in selected_labels_dict')
        nodes_selected.extend(random.sample(nodes_in_each_cluster[cluster], count))

    H = getInducedSubgraph(G, nodes_selected)
    partition = debug.louvain_partitions(H, seed=0,resolution=res)
    partition_ = deque(partition, maxlen=1).pop()
    H_label = part_to_full_label(partition_,G.number_of_nodes())
    label_compressed, H_label_compressed = get_compressed_labels(label, H_label, nodes_selected)
    NMI_ = NMI(H_label_compressed, label_compressed)
    Purity_ = met.purity_score(label_compressed, H_label_compressed)
    return NMI_, Purity_


'''
label is in the form of [-1,0,1,4,2,3,1,1,1,...] / -1 meaning not assigned with cluster (Ignore)
partition is in the form of [{1,2,3}, {4},...] / list of sets
'''
def label_to_partition(label, only_labeled_nodes = 1):
    partition = defaultdict(set)
    singleton_idx = len(label)+1
    if only_labeled_nodes == 1: #Case where -1 is ignored
        for i, node_label in enumerate(label):
            if node_label != -1:
                partition[node_label].add(i)
    elif only_labeled_nodes == 0: #Case where -1 is partitioned as singleton 
        for i, node_label in enumerate(label):
            if node_label == -1:
                partition[singleton_idx].add(i)
                singleton_idx += 1
            else:
                partition[node_label].add(i)

    return list(partition.values())

'''
Does the Louvain and return the partition and label
'''
def do_Louvain(Graph_to_part, G, res, seed, label):
    partition = debug.louvain_partitions(Graph_to_part, seed=seed,resolution=res)
    partition_ = deque(partition, maxlen=1).pop()
    label_ = part_to_full_label(partition_,G.number_of_nodes())
    NMI_ = NMI(label_, label)
    Purity_ = met.purity_score(label, label_)
    return label_, NMI_, Purity_

def do_Leiden(Graph_to_part, G, res, seed, label):
    G_ig, mapping, _ = networkx_to_igraph(Graph_to_part) # Mapping = igraph index to networkx index
    partition = la.find_partition(G_ig, la.RBConfigurationVertexPartition, n_iterations=-1, resolution_parameter=res, seed=seed)
    label_ = part_to_full_label(partition, G.number_of_nodes(), mapping)
    NMI_ = NMI(label_, label)
    Purity_ = met.purity_score(label, label_)
    return label_, NMI_, Purity_

def do_random_walks(G,unselected_nodes, label, H_label, node_to_FR, walk_len = 5, walk_rep = 20):
    #precalculate the list of neighbor nodes with higher FR values
    neighbor_nodes_dict = defaultdict(list)
    for node in G.nodes():
        neighbor_nodes = [i for i in G.neighbors(node) if node_to_FR[i] > node_to_FR[node]]
        neighbor_nodes_dict[node] = neighbor_nodes

    #Only refer to labelings of H before the random walk    
    new_label = H_label.copy()

    for node in unselected_nodes:
        visited_label_cntr = defaultdict(int)
        for _ in range(walk_rep):
            next_node = node
            for _ in range(walk_len):
                '''
                Do we do random walk on original G with possibility of ending up on unlabeled nodes?
                Only count the final node visited or some weighted sum of all nodes visited?
                '''
                #randomly pick a neighbor node
                if len(neighbor_nodes_dict[next_node]) == 0:
                    break
                next_node = random.choice(neighbor_nodes_dict[next_node])
            visited_label_cntr[H_label[next_node]] += 1
            # print('node:' , node, 'visited node:' , next_node, 'visited label:' , H_label[next_node])
        #Assign the node to the most visited label
        new_label[node] = max(visited_label_cntr, key=visited_label_cntr.get)
    
    NMI_, Purity_ = calc_NMI_Purity(label, new_label)
    return new_label, NMI_, Purity_
            
                


def plot_to_pdf(resolution_list, data_for_plot, k, True_num_of_clusters, pdf_name, data_name):
    plt.figure(figsize=(15,15)) # figure for NMI vs purity
    plt.figure(figsize=(15,8)) # figure for NMI vs Purity (whole)
    # plt.figure(figsize=(15,15)) # figure for NMI
    # plt.figure(figsize=(15,15)) # figure for Purity
    # plt.figure(figsize=(15,15)) # figure for preserve_ratio
    # plt.figure(figsize=(15,15)) # figure for balancedness
    # plt.figure(figsize=(10,15)) # figure for InEdge Degrees

    FR_tp = ['FL','FL_ng','FL_ng_pr','PageRank']
    #NMI_vs_Purity, NMI_vs_Purity_whole, nmi_fig, purity_fig, preserve_fig, balance_fig, InEdge_fig = plt.get_fignums()[-7:]
    NMI_vs_Purity, NMI_vs_Purity_whole = plt.get_fignums()[-2:]
    max_dot_size = 100
    #plot NMI vs Purity
    markers = ['o','s','*','^','h','D'] #idx (resolution)
    colors = ['b','g','y','r'] #FR_type
    
    for idx, res in enumerate(resolution_list):
        for FR_type in range(4):
            data = data_for_plot[idx][FR_type]

            plt.figure(NMI_vs_Purity)
            plt.subplot(3,2,idx+1)
            plt.scatter(data['Purity_after_vote'], data['NMI_after_vote'], color = colors[FR_type], s = data['percent_nodes']/100*max_dot_size,label = str(FR_tp[FR_type])+ '|'+str(round(data['percent_nodes'],2))+ '%' + '|#Com:'+str(data['Num_Cluster_Louv_on_H']))
            #if data_for_plot has key 'extra_name' then plot
            if data.get('extra_name') is not None:
                plt.scatter(data['extra_purity'], data['extra_nmi'], edgecolors = colors[FR_type], facecolors = 'none', s = data['extra_percent_nodes']/100*max_dot_size,label = data['extra_name']+ '|'+str(FR_tp[FR_type])+ '|'+str(round(data['extra_percent_nodes'],2))+ '%' + '|#Com:'+str(data['Num_Cluster_Louv_on_H']))
            if FR_type==0:
                plt.scatter(data['Louvain_Purity'], data['Louvain_NMI'], color = 'k', s = max_dot_size,label = 'Louv_baseline'+ '|#Com:'+str(data['Num_Cluster_Louv_on_G']))
                plt.scatter(data['Leiden_Purity'], data['Leiden_NMI'], color = 'm', s = max_dot_size,label = 'Leiden_baseline'+ '|#Com:'+str(data['Num_Cluster_Leiden_on_G']))

            plt.figure(NMI_vs_Purity_whole)
            plt.scatter(data['Purity_after_vote'], data['NMI_after_vote'], color = colors[FR_type], marker = markers[idx] , s = data['percent_nodes']/100*max_dot_size,label = str(FR_tp[FR_type])+ '|'+str(round(data['percent_nodes'],2))+ '%' + '|#Com:'+str(data['Num_Cluster_Louv_on_H']))
            if data.get('extra_name') is not None:
                plt.scatter(data['extra_purity'], data['extra_nmi'], edgecolors = colors[FR_type], facecolors = 'none', s = data['extra_percent_nodes']/100*max_dot_size,label = data['extra_name']+ '|'+str(FR_tp[FR_type])+ '|'+str(round(data['extra_percent_nodes'],2))+ '%' + '|#Com:'+str(data['Num_Cluster_Louv_on_H']))
            if FR_type==0:
                plt.scatter(data['Louvain_Purity'], data['Louvain_NMI'], color = 'k', s = max_dot_size, marker = markers[idx], label = 'Louv_baseline|'+str(res))
                plt.scatter(data['Leiden_Purity'], data['Leiden_NMI'], color = 'm', s = max_dot_size, marker = markers[idx], label = 'Leiden_baseline|'+str(res))
    
        plt.figure(NMI_vs_Purity)
        plt.tight_layout(rect=[0,0,1,1])
        plt.legend(loc = 'center left', bbox_to_anchor=(1, 0.5), fontsize=9)
        plt.subplot(3,2,idx+1)
        plt.xlabel('Purity')
        plt.ylabel('NMI')
        plt.title(data_name + ' | top ' + str(k*100) + '%' + '| Num_hops: log(n) |res: ' + str(res)+ '|#OfComm: '+str(True_num_of_clusters))
        
        plt.figure(NMI_vs_Purity_whole)
        plt.xlabel('Purity')
        plt.ylabel('NMI')
        plt.title(data_name + ' | top ' + str(k*100) + '%' + '| Num_hops: log(n)')
        plt.legend(loc='center left', bbox_to_anchor=(1, 0.5), fontsize=11, ncol=2)
        plt.tight_layout(rect=[0, 0, 1, 1])
    
    p = PdfPages(pdf_name)            
    fig_nums = plt.get_fignums() 
    figs = [plt.figure(n) for n in fig_nums] 
	
    for fig in figs: 
        fig.savefig(p, format='pdf') 
    p.close() 
    plt.close('all')

def write_to_excel (data, excel_name, sheet_name):
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
    from openpyxl.styles import Font

    row = data['data_idx']*17+2
    col = data['method_idx']*6+3

    wb = load_workbook(excel_name)
    ws = wb[sheet_name]

    
    '''
    write down method name
    '''
    ws.merge_cells(start_row = 1, start_column = col, end_row = 1, end_column = col+5)
    ws.cell(row=1, column=col, value=data['method_name']).font = Font(bold=True)
    
    ws.cell(row=1, column=col).alignment = Alignment(horizontal='center')
    ws.cell(row = row, column = 1, value = data['data_name']).font = Font(bold=True)
    '''
    headers
    '''
    headers = ['NMI', 'Purity', '% nodes', '# of Comm'+ '|'+ str(data['num_comm'])]
    '''
    Write the results
    '''
    for res_idx, res in enumerate(data['res_list']):
        ws.cell(row = row, column = col+res_idx, value = res).font = Font(bold=True)
        ws.cell(row= row, column = 2, value = 'Resolution').font = Font(bold=True)
        for fr_idx, fr in enumerate(data['fr_tp']):
            #nmi, purity, percent_nodes, num_comm = data[res][fr]
            lis = data[res][fr]
            for i in range(4):
                ws.cell(row = row+1+i+4*fr_idx, column = col+res_idx, value = lis[i])
                #Write the headers
                ws.cell(row = row+1+i+4*fr_idx, column = 2, value = headers[i]).font = Font(bold=True)
                if i == 0:
                    ws.cell(row = row+1+i+4*fr_idx, column = 1, value = fr).font = Font(bold=True)
    '''
    onto seprate sheet
    '''
    if data['data_name'] in wb.sheetnames:
        ws = wb[data['data_name']]
    else:
        ws = wb.create_sheet(title=data['data_name'])
    
    row = 2
    '''
    write down method name
    '''
    ws.merge_cells(start_row = 1, start_column = col, end_row = 1, end_column = col+5)
    ws.cell(row=1, column=col, value=data['method_name']+ '|#Comm:' + str(data['num_comm'])).font = Font(bold=True)
    ws.cell(row=1, column=col).alignment = Alignment(horizontal='center')
    ws.cell(row = row, column = 1, value = data['data_name']).font = Font(bold=True)

    '''
    Write the results
    '''
    for res_idx, res in enumerate(data['res_list']):
        ws.cell(row = row, column = col+res_idx, value = res).font = Font(bold=True)
        ws.cell(row= row, column = 2, value = 'Resolution').font = Font(bold=True)
        for fr_idx, fr in enumerate(data['fr_tp']):
            #nmi, purity, percent_nodes, num_comm = data[res][fr]
            lis = data[res][fr]
            for i in range(4):
                ws.cell(row = row+1+i+4*fr_idx, column = col+res_idx, value = lis[i])
                #Write the headers
                ws.cell(row = row+1+i+4*fr_idx, column = 2, value = headers[i]).font = Font(bold=True)
                if i == 0:
                    ws.cell(row = row+1+i+4*fr_idx, column = 1, value = fr).font = Font(bold=True)
    wb.save(excel_name)


def freeze_panes(excel_name):
    import openpyxl
    from openpyxl import load_workbook

    wb = load_workbook(excel_name)
    for sheet in wb.worksheets:
        #freeze first two columns
        sheet.freeze_panes = 'C3'
    wb.save(excel_name)